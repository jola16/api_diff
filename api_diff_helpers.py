"""Helper functions for api_diff."""

import logging
from http import HTTPStatus
import sys
from typing import Any
import requests
from openpyxl.worksheet.table import Table, TableStyleInfo
import openpyxl

logger = logging.getLogger(__name__)


class RateLimitError(Exception):
    """Exception for rate limit."""

    def __init__(self) -> None:
        """Initialize the exception with rate limit message."""
        super().__init__("Rate limit")


def save_to_excel(results: list[dict[str, Any]], filename: str = "api_diff_results.xlsx") -> None:
    """Save the results to an Excel file."""
    if not results:
        return
    # Create Excel document
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "API Diff Results"

    # Write headers
    headers = list(results[0].keys())
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header.replace("_", " ").title())

    # Write data
    for row_num, result in enumerate(results, 2):
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=row_num, column=col_num, value=result[header])

    # Autosize columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Create table
    last_col = chr(ord("A") + len(headers) - 1)
    tab = Table(
        displayName="APIDiffTable",
        ref=f"A1:{last_col}{len(results)+1}",
    )
    style = TableStyleInfo(name="TableStyleLight14", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Save
    wb.save(filename)
    logger.info("Saved results to %s", filename)


def fetch(
    base_url: str,
    *,
    method: str = "GET",
    params: dict[str, Any] | None = None,
    headers: dict[str, str] | None = None,
) -> dict[str, Any]:
    """Fetch data from API.

    Args:
        base_url: The base URL of the API endpoint.
        method: HTTP method to use (default: "GET").
        params: Query parameters for GET or JSON body for other methods (default: None).
        headers: HTTP headers to include (default: None).

    Returns:
        The JSON response from the API.
    """
    if params is None:
        params = {}
    if headers is None:
        headers = {}
    try:
        if method.upper() == "GET":
            response: requests.Response = requests.get(
                base_url, headers=headers, params=params, timeout=10
            )
        else:
            response: requests.Response = requests.request(
                method.upper(), base_url, headers=headers, json=params, timeout=10
            )
        logger.debug("Fetching %s with params %s", base_url, params)
        status_phrase = HTTPStatus(response.status_code).phrase
        logger.debug("Status: %s %s", response.status_code, status_phrase)
        if response.status_code == HTTPStatus.TOO_MANY_REQUESTS.value:
            logger.warning("Rate limit hit, retrying...")
            raise RateLimitError
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException:
        logger.exception("Request failed for %s with params %s", base_url, params)
        sys.exit(1)
    except ValueError:
        logger.exception("JSON decode failed for %s with params %s", base_url, params)
        logger.exception("Response text: %s", response.text[:500])
        sys.exit(1)
